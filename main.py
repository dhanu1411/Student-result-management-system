from openpyxl import Workbook, load_workbook
import os

# Function to validate marks
def get_marks(subject):

    while True:

        try:
            marks = int(input(f"{subject} marks: "))

            if 0 <= marks <= 100:
                return marks

            print("Invalid marks! Enter marks between 0 and 100.")

        except ValueError:
            print("Please enter a valid number.")


filename = "student_results.xlsx"

# Check if file exists
if os.path.exists(filename):

    workbook = load_workbook(filename)
    sheet = workbook.active

    headers = [cell.value for cell in sheet[1]]

    # Add Status column if missing
    if "Status" not in headers:

        sheet.cell(row=1, column=len(headers) + 1).value = "Status"

        headers = [cell.value for cell in sheet[1]]

    # Update old records with Pass/Fail status
    status_col = headers.index("Status") + 1

    for row in range(2, sheet.max_row + 1):

        if sheet.cell(row=row, column=status_col).value is None:

            tamil = sheet.cell(row=row, column=2).value
            english = sheet.cell(row=row, column=3).value
            maths = sheet.cell(row=row, column=4).value
            science = sheet.cell(row=row, column=5).value

            if tamil < 35 or english < 35 or maths < 35 or science < 35:
                sheet.cell(row=row, column=status_col).value = "Fail"
            else:
                sheet.cell(row=row, column=status_col).value = "Pass"

else:

    workbook = Workbook()
    sheet = workbook.active

    sheet.append([
        "Name",
        "Tamil",
        "English",
        "Maths",
        "Science",
        "Total",
        "Average",
        "Grade",
        "Status"
    ])

# Number of students
num = int(input("Enter number of students: "))

for i in range(num):

    print(f"\nStudent {i + 1}")

    name = input("Enter student name: ")

    tamil = get_marks("Tamil")
    english = get_marks("English")
    maths = get_marks("Maths")
    science = get_marks("Science")

    total = tamil + english + maths + science
    average = total / 4

    # Grade Calculation
    if average >= 90:
        grade = "A+"
    elif average >= 75:
        grade = "A"
    elif average >= 60:
        grade = "B"
    elif average >= 40:
        grade = "C"
    else:
        grade = "Fail"

    # Pass / Fail Status
    if tamil < 35 or english < 35 or maths < 35 or science < 35:
        status = "Fail"
    else:
        status = "Pass"

    # Add data to Excel
    sheet.append([
        name,
        tamil,
        english,
        maths,
        science,
        total,
        average,
        grade,
        status
    ])

# Save Excel file
workbook.save(filename)

print("\nData added successfully!")